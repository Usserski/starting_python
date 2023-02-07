import openpyxl as xl
from openpyxl import Bart


wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_prize = cell.value * 0.9
    corrected_prize_cell = sheet.cell(row, 4)
    corrected_prize_cell.value = corrected_prize

wb.save('transactions2.xlsx')
